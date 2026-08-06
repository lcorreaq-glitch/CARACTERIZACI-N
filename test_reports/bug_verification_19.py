import io
import json
import os
from pathlib import Path

import openpyxl
import requests
from pymongo import MongoClient


ROOT = Path("/app")
FRONT_ENV = (ROOT / "frontend" / ".env").read_text()
BACK_ENV = (ROOT / "backend" / ".env").read_text()
APP_LAYOUT = ROOT / "frontend" / "src" / "pages" / "AppLayout.jsx"


def env_value(text, key, default=None):
    for line in text.splitlines():
        if line.startswith(key + "="):
            return line.split("=", 1)[1].strip().strip('"')
    return default


BASE = os.environ.get("TEST_BASE_URL") or env_value(FRONT_ENV, "REACT_APP_BACKEND_URL", "http://localhost:8001")
API = BASE.rstrip("/") + "/api"
MONGO_URL = env_value(BACK_ENV, "MONGO_URL", "mongodb://localhost:27017")
DB_NAME = env_value(BACK_ENV, "DB_NAME", "iudigital")

GRUPO_FCEAC_IN_SCOPE = "PREEAC2602B010224"
GRUPO_TRABAJO_SOCIAL_OUT_SCOPE = "PRECBH2602B020258"
EXPECTED_VISTA_COLUMNS = [
    "Cedula",
    "Nombre",
    "Apellidos",
    "Programa",
    "Promedio",
    "Estado matricula",
    "Flags de vulnerabilidad",
    "N flags",
]


def login(email, password):
    return requests.post(f"{API}/auth/login", json={"email": email, "password": password}, timeout=30)


def headers(token):
    return {"Authorization": f"Bearer {token}"}


def workbook_from_response(resp):
    return openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)


def worksheet_rows(ws):
    values = list(ws.iter_rows(values_only=True))
    header = list(values[0]) if values else []
    rows = [dict(zip(header, row)) for row in values[1:]]
    return header, rows


def ui_flags(student):
    flags = []
    if student.get("grupo_vulnerable"):
        flags.append(student.get("tipo_grupo_vulnerable") or "Vulnerable")
    if student.get("victima_conflicto"):
        flags.append("Víctima del conflicto")
    if student.get("discapacidad_flag"):
        flags.append(f"Discapacidad: {student.get('discapacidad_tipo') or 'No especificada'}")
    if student.get("sisben_tiene") and student.get("sisben_nivel"):
        flags.append(f"SISBEN {student.get('sisben_nivel')}")
    if student.get("tipo_ubicacion") in ("Rural", "Semirural"):
        flags.append(student.get("tipo_ubicacion"))
    if student.get("etnia") and student.get("etnia") not in ("Ninguno", "No Aplica"):
        flags.append(f"Etnia: {student.get('etnia')}")
    return " | ".join(flags) if flags else "Sin flags", len(flags)


def get_json_or_text(resp):
    try:
        return resp.json()
    except Exception:
        return resp.text[:300]


def find_docente_group(db, docente_id):
    if not docente_id:
        return None
    for group in db.grupos.find({"docente_id": docente_id}, {"_id": 0, "codigo_grupo": 1}).limit(500):
        codigo = group.get("codigo_grupo")
        if codigo and db.matriculas.count_documents({"codigo_grupo": codigo}, limit=1):
            return codigo
    group = db.grupos.find_one({"docente_id": docente_id}, {"_id": 0, "codigo_grupo": 1})
    return (group or {}).get("codigo_grupo")


def main():
    results = {
        "api_base": API,
        "skill_lookup": "No relevant testing skill found.",
        "checks": [],
        "failures": [],
        "notes": [],
        "groups": {},
        "restored": [],
    }

    def check(name, condition, evidence=None):
        item = {"name": name, "passed": bool(condition), "evidence": evidence}
        results["checks"].append(item)
        if not condition:
            results["failures"].append(item)

    client = MongoClient(MONGO_URL)
    db = client[DB_NAME]
    original_user_download = {}
    original_global_setting = None
    global_setting_existed = False

    try:
        # Source-level confirmation for the global filters panel: Materia/materia_id is no longer an accordion filter.
        app_layout = APP_LAYOUT.read_text()
        filter_panel_section = app_layout.split("function FiltersPanel()", 1)[1].split("function DocenteCursosPanel()", 1)[0]
        check(
            "AppLayout FiltersPanel has no Materia/materia_id global filter trigger",
            "filter-materia" not in filter_panel_section
            and "materia_id" not in filter_panel_section
            and 'key: "materia' not in filter_panel_section,
            {"file": str(APP_LAYOUT), "filter_materia_occurrences": filter_panel_section.count("materia")},
        )
        check(
            "AppLayout global object filters contain Docente and Grupo only for docente/materia/grupo area",
            'key: "docente_id"' in filter_panel_section and 'key: "codigo_grupo"' in filter_panel_section,
            {"file": str(APP_LAYOUT)},
        )

        # DB sanity for exact requested scope groups.
        for label, codigo in [("in_scope_fceac", GRUPO_FCEAC_IN_SCOPE), ("out_scope_trabajo_social", GRUPO_TRABAJO_SOCIAL_OUT_SCOPE)]:
            group = db.grupos.find_one({"codigo_grupo": codigo}, {"_id": 0, "codigo_grupo": 1, "facultad": 1, "programa": 1, "docente_id": 1})
            matriculas = db.matriculas.count_documents({"codigo_grupo": codigo})
            results["groups"][label] = {"group": group, "matriculas": matriculas}
            check(f"test group {codigo} exists with matriculas", bool(group) and matriculas > 0, {"group": group, "matriculas": matriculas})

        # Superadmin: unrestricted download and workbook contract.
        super_login = login("lcorreaq@gmail.com", "IUDigital2026")
        check("superadmin login", super_login.status_code == 200, {"status": super_login.status_code, "body": super_login.text[:200]})
        super_token = super_login.json()["access_token"] if super_login.status_code == 200 else None

        if super_token:
            export = requests.get(f"{API}/exports/grupo/{GRUPO_FCEAC_IN_SCOPE}/vista?fmt=xlsx", headers=headers(super_token), timeout=60)
            check(
                "superadmin can download in-scope sample vista export",
                export.status_code == 200,
                {"status": export.status_code, "content_type": export.headers.get("content-type"), "bytes": len(export.content)},
            )
            if export.status_code == 200:
                wb = workbook_from_response(export)
                check("vista xlsx has exactly two sheets", wb.sheetnames == ["Grupo", "Vista docente"], {"sheetnames": wb.sheetnames})
                header, rows = worksheet_rows(wb["Vista docente"])
                check("Vista docente sheet has exactly the 8 requested columns", header == EXPECTED_VISTA_COLUMNS, {"header": header})
                check("Vista docente export has student rows", len(rows) > 0, {"row_count": len(rows)})
                extra_columns = [c for c in header if c not in EXPECTED_VISTA_COLUMNS]
                check("Vista docente sheet contains no non-visible/full-characterization columns", not extra_columns and len(header) == 8, {"extra_columns": extra_columns})

                detail = requests.get(f"{API}/admin/grupos/{GRUPO_FCEAC_IN_SCOPE}", headers=headers(super_token), timeout=30)
                check("group detail API available for visible-data comparison", detail.status_code == 200, {"status": detail.status_code})
                if detail.status_code == 200:
                    detail_json = detail.json()
                    students = detail_json.get("estudiantes", [])
                    check(
                        "export row count matches group-detail visible matriculados",
                        len(rows) == len(students) == detail_json.get("total_estudiantes"),
                        {"export_rows": len(rows), "detail_rows": len(students), "detail_total": detail_json.get("total_estudiantes")},
                    )
                    rows_by_cedula = {str(row.get("Cedula")): row for row in rows}
                    mismatches = []
                    for student in students:
                        cedula = str(student.get("cedula"))
                        expected_flags, expected_n = ui_flags(student)
                        actual = rows_by_cedula.get(cedula)
                        if not actual:
                            mismatches.append({"cedula": cedula, "issue": "missing export row"})
                            continue
                        actual_n = int(actual.get("N flags") or 0)
                        if actual.get("Flags de vulnerabilidad") != expected_flags or actual_n != expected_n:
                            mismatches.append({
                                "cedula": cedula,
                                "expected_flags": expected_flags,
                                "expected_n": expected_n,
                                "actual_flags": actual.get("Flags de vulnerabilidad"),
                                "actual_n": actual_n,
                            })
                    check(
                        "vista export flags match the same visible red-flag logic as group detail",
                        len(mismatches) == 0,
                        {"mismatch_count": len(mismatches), "first_mismatches": mismatches[:3]},
                    )

            super_any = requests.get(f"{API}/exports/grupo/{GRUPO_TRABAJO_SOCIAL_OUT_SCOPE}/vista?fmt=xlsx", headers=headers(super_token), timeout=60)
            check(
                "superadmin can download any vista group including Trabajo Social sample",
                super_any.status_code == 200,
                {"status": super_any.status_code, "bytes": len(super_any.content)},
            )

        # Force relevant permission state temporarily, then restore exactly.
        users_to_touch = ["decano.test@iudigital.edu.co", "coord.test@iudigital.edu.co", "docente.demo@iudigital.edu.co"]
        for email in users_to_touch:
            u = db.users.find_one({"email": email}, {"_id": 0, "id": 1, "download_enabled": 1})
            check(f"user exists for permission setup: {email}", bool(u), u)
            if u:
                original_user_download[email] = u.get("download_enabled", None)
        if "decano.test@iudigital.edu.co" in original_user_download:
            db.users.update_one({"email": "decano.test@iudigital.edu.co"}, {"$set": {"download_enabled": True}})
        if "coord.test@iudigital.edu.co" in original_user_download:
            db.users.update_one({"email": "coord.test@iudigital.edu.co"}, {"$set": {"download_enabled": True}})
        if "docente.demo@iudigital.edu.co" in original_user_download:
            db.users.update_one({"email": "docente.demo@iudigital.edu.co"}, {"$set": {"download_enabled": False}})

        settings = db.system_settings.find_one({"_id": "global"}, {"_id": 0})
        global_setting_existed = bool(settings)
        original_global_setting = (settings or {}).get("docente_downloads_globally_enabled", None)
        db.system_settings.update_one({"_id": "global"}, {"$set": {"docente_downloads_globally_enabled": False}}, upsert=True)

        decano_user = db.users.find_one({"email": "decano.test@iudigital.edu.co"}, {"_id": 0})
        decano_fac = None
        if decano_user and decano_user.get("facultad_id"):
            fac = db.facultades.find_one({"id": decano_user["facultad_id"]}, {"_id": 0, "nombre": 1})
            decano_fac = (fac or {}).get("nombre")
        check("decano.test is scoped to FCEAC/facultad and has download_enabled=true for test", bool(decano_fac) and decano_user.get("download_enabled") is True, {"facultad": decano_fac, "download_enabled": decano_user.get("download_enabled") if decano_user else None})

        decano_login = login("decano.test@iudigital.edu.co", "Decano2026!")
        check("decano.test login", decano_login.status_code == 200, {"status": decano_login.status_code, "body": decano_login.text[:200]})
        if decano_login.status_code == 200:
            decano_token = decano_login.json()["access_token"]
            out_resp = requests.get(f"{API}/exports/grupo/{GRUPO_TRABAJO_SOCIAL_OUT_SCOPE}/vista?fmt=xlsx", headers=headers(decano_token), timeout=30)
            in_resp = requests.get(f"{API}/exports/grupo/{GRUPO_FCEAC_IN_SCOPE}/vista?fmt=xlsx", headers=headers(decano_token), timeout=60)
            check(
                "CRITICAL: decano.test with download_enabled=true receives 403 for Trabajo Social/FCYH group",
                out_resp.status_code == 403,
                {"status": out_resp.status_code, "body": get_json_or_text(out_resp), "group": GRUPO_TRABAJO_SOCIAL_OUT_SCOPE},
            )
            check(
                "decano.test with download_enabled=true receives 200 for FCEAC group",
                in_resp.status_code == 200,
                {"status": in_resp.status_code, "content_type": in_resp.headers.get("content-type"), "bytes": len(in_resp.content), "group": GRUPO_FCEAC_IN_SCOPE},
            )

        coord_user = db.users.find_one({"email": "coord.test@iudigital.edu.co"}, {"_id": 0})
        coord_prog = None
        if coord_user and coord_user.get("programa_id"):
            prog = db.programas.find_one({"id": coord_user["programa_id"]}, {"_id": 0, "nombre": 1})
            coord_prog = (prog or {}).get("nombre")
        check("coord.test is scoped to Administración de Empresas and has download_enabled=true for test", bool(coord_prog) and coord_user.get("download_enabled") is True, {"programa": coord_prog, "download_enabled": coord_user.get("download_enabled") if coord_user else None})
        coord_login = login("coord.test@iudigital.edu.co", "Coord2026!")
        check("coord.test login", coord_login.status_code == 200, {"status": coord_login.status_code, "body": coord_login.text[:200]})
        if coord_login.status_code == 200:
            coord_token = coord_login.json()["access_token"]
            coord_out = requests.get(f"{API}/exports/grupo/{GRUPO_TRABAJO_SOCIAL_OUT_SCOPE}/vista?fmt=xlsx", headers=headers(coord_token), timeout=30)
            coord_in = requests.get(f"{API}/exports/grupo/{GRUPO_FCEAC_IN_SCOPE}/vista?fmt=xlsx", headers=headers(coord_token), timeout=60)
            check(
                "coordinador with download_enabled=true receives 403 for non-program Trabajo Social group",
                coord_out.status_code == 403,
                {"status": coord_out.status_code, "body": get_json_or_text(coord_out), "group": GRUPO_TRABAJO_SOCIAL_OUT_SCOPE},
            )
            check(
                "coordinador with download_enabled=true receives 200 for Administración de Empresas group",
                coord_in.status_code == 200,
                {"status": coord_in.status_code, "content_type": coord_in.headers.get("content-type"), "bytes": len(coord_in.content), "group": GRUPO_FCEAC_IN_SCOPE},
            )

        docente_user = db.users.find_one({"email": "docente.demo@iudigital.edu.co"}, {"_id": 0, "id": 1, "download_enabled": 1, "role": 1})
        docente_group = find_docente_group(db, (docente_user or {}).get("id")) or GRUPO_FCEAC_IN_SCOPE
        docente_login = login("docente.demo@iudigital.edu.co", "Docente2026")
        check("docente.demo login", docente_login.status_code == 200, {"status": docente_login.status_code, "body": docente_login.text[:200]})
        if docente_login.status_code == 200:
            docente_token = docente_login.json()["access_token"]
            docente_own = requests.get(f"{API}/exports/grupo/{docente_group}/vista?fmt=xlsx", headers=headers(docente_token), timeout=30)
            docente_other = requests.get(f"{API}/exports/grupo/{GRUPO_FCEAC_IN_SCOPE}/vista?fmt=xlsx", headers=headers(docente_token), timeout=30)
            check(
                "docente without download_enabled receives 403 on own/assigned vista download",
                docente_user.get("download_enabled") is False and docente_own.status_code == 403,
                {"status": docente_own.status_code, "body": get_json_or_text(docente_own), "tested_group": docente_group, "global_download": False},
            )
            check(
                "docente without download_enabled receives 403 on another vista download",
                docente_other.status_code == 403,
                {"status": docente_other.status_code, "body": get_json_or_text(docente_other), "tested_group": GRUPO_FCEAC_IN_SCOPE},
            )

            # Positive teacher flow for the user request: an individually enabled docente can download
            # the view-only export for their own group, but not another docente's group.
            db.users.update_one({"email": "docente.demo@iudigital.edu.co"}, {"$set": {"download_enabled": True}})
            docente_enabled_own = requests.get(f"{API}/exports/grupo/{docente_group}/vista?fmt=xlsx", headers=headers(docente_token), timeout=60)
            docente_enabled_other = requests.get(f"{API}/exports/grupo/{GRUPO_FCEAC_IN_SCOPE}/vista?fmt=xlsx", headers=headers(docente_token), timeout=30)
            check(
                "enabled docente can download vista xlsx for own group",
                docente_enabled_own.status_code == 200,
                {"status": docente_enabled_own.status_code, "content_type": docente_enabled_own.headers.get("content-type"), "bytes": len(docente_enabled_own.content), "tested_group": docente_group},
            )
            if docente_enabled_own.status_code == 200:
                wb_doc = workbook_from_response(docente_enabled_own)
                header_doc, rows_doc = worksheet_rows(wb_doc["Vista docente"])
                check(
                    "enabled docente own-group vista xlsx keeps 2 sheets and 8 requested columns",
                    wb_doc.sheetnames == ["Grupo", "Vista docente"] and header_doc == EXPECTED_VISTA_COLUMNS and len(rows_doc) > 0,
                    {"sheetnames": wb_doc.sheetnames, "header": header_doc, "row_count": len(rows_doc)},
                )
            check(
                "enabled docente still cannot download another docente's vista group",
                docente_enabled_other.status_code == 403,
                {"status": docente_enabled_other.status_code, "body": get_json_or_text(docente_enabled_other), "tested_group": GRUPO_FCEAC_IN_SCOPE},
            )

    finally:
        for email, original in original_user_download.items():
            if original is None:
                db.users.update_one({"email": email}, {"$unset": {"download_enabled": ""}})
            else:
                db.users.update_one({"email": email}, {"$set": {"download_enabled": original}})
            results["restored"].append({"email": email, "download_enabled": original})
        if original_global_setting is None:
            if global_setting_existed:
                db.system_settings.update_one({"_id": "global"}, {"$unset": {"docente_downloads_globally_enabled": ""}})
            else:
                # Preserve any other fields created by app code, but remove the test-created toggle.
                db.system_settings.update_one({"_id": "global"}, {"$unset": {"docente_downloads_globally_enabled": ""}})
        else:
            db.system_settings.update_one({"_id": "global"}, {"$set": {"docente_downloads_globally_enabled": original_global_setting}}, upsert=True)
        results["restored"].append({"system_settings.docente_downloads_globally_enabled": original_global_setting})
        client.close()

    results_path = ROOT / "test_reports" / "bug_verification_19_results.json"
    results_path.write_text(json.dumps(results, ensure_ascii=False, indent=2, default=str))
    print(json.dumps(results, ensure_ascii=False, indent=2, default=str))
    return 1 if results["failures"] else 0


if __name__ == "__main__":
    raise SystemExit(main())