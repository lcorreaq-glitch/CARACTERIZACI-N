import io
import json
import os
from pathlib import Path

import openpyxl
import requests
from pymongo import MongoClient


ROOT = Path("/app")
FRONT_ENV = (ROOT / "frontend" / ".env").read_text()
BACK_ENV = (ROOT / "backend" / ".env").read_text()


def env_value(text, key, default=None):
    for line in text.splitlines():
        if line.startswith(key + "="):
            return line.split("=", 1)[1].strip().strip('"')
    return default


BASE = os.environ.get("TEST_BASE_URL") or env_value(FRONT_ENV, "REACT_APP_BACKEND_URL", "http://localhost:8001")
API = BASE.rstrip("/") + "/api"
MONGO_URL = env_value(BACK_ENV, "MONGO_URL", "mongodb://localhost:27017")
DB_NAME = env_value(BACK_ENV, "DB_NAME", "iudigital")
GRUPO_TARGET = "PREEAC2602B010224"


def login(email, password):
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": password}, timeout=30)
    return r


def headers(token):
    return {"Authorization": f"Bearer {token}"}


def ui_flags(s):
    flags = []
    if s.get("grupo_vulnerable"):
        flags.append(s.get("tipo_grupo_vulnerable") or "Vulnerable")
    if s.get("victima_conflicto"):
        flags.append("Víctima del conflicto")
    if s.get("discapacidad_flag"):
        flags.append(f"Discapacidad: {s.get('discapacidad_tipo') or 'No especificada'}")
    if s.get("sisben_tiene") and s.get("sisben_nivel"):
        flags.append(f"SISBEN {s.get('sisben_nivel')}")
    if s.get("tipo_ubicacion") in ("Rural", "Semirural"):
        flags.append(s.get("tipo_ubicacion"))
    if s.get("etnia") and s.get("etnia") not in ("Ninguno", "No Aplica"):
        flags.append(f"Etnia: {s.get('etnia')}")
    return " | ".join(flags) if flags else "Sin flags", len(flags)


def workbook_from_response(resp):
    return openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)


def rows_as_dicts(ws):
    values = list(ws.iter_rows(values_only=True))
    header = list(values[0]) if values else []
    return header, [dict(zip(header, row)) for row in values[1:]]


def find_group_with_students(db, query):
    for g in db.grupos.find(query, {"_id": 0, "codigo_grupo": 1, "facultad": 1, "programa": 1}).limit(500):
        if db.matriculas.count_documents({"codigo_grupo": g["codigo_grupo"]}, limit=1):
            return g
    return db.grupos.find_one(query, {"_id": 0, "codigo_grupo": 1, "facultad": 1, "programa": 1})


def norm(value):
    return (value or "").strip().lower()


def find_group_by_predicate(db, predicate):
    fallback = None
    for g in db.grupos.find({}, {"_id": 0, "codigo_grupo": 1, "facultad": 1, "programa": 1}).limit(5000):
        if predicate(g):
            if fallback is None:
                fallback = g
            if db.matriculas.count_documents({"codigo_grupo": g["codigo_grupo"]}, limit=1):
                return g
    return fallback


def main():
    results = {
        "api_base": API,
        "checks": [],
        "failures": [],
        "notes": [],
        "scope_test_groups": {},
        "restored_users": [],
    }
    client = MongoClient(MONGO_URL)
    db = client[DB_NAME]
    original_user_values = {}

    def check(name, condition, evidence=None):
        item = {"name": name, "passed": bool(condition), "evidence": evidence}
        results["checks"].append(item)
        if not condition:
            results["failures"].append(item)

    try:
        # Superadmin export: workbook contract and data visible in group detail.
        super_login = login("lcorreaq@gmail.com", "IUDigital2026")
        check("superadmin login", super_login.status_code == 200, {"status": super_login.status_code, "body": super_login.text[:200]})
        super_token = super_login.json()["access_token"]

        export_resp = requests.get(f"{API}/exports/grupo/{GRUPO_TARGET}/vista?fmt=xlsx", headers=headers(super_token), timeout=60)
        check("superadmin vista export status 200", export_resp.status_code == 200, {"status": export_resp.status_code, "content_type": export_resp.headers.get("content-type")})
        if export_resp.status_code == 200:
            wb = workbook_from_response(export_resp)
            check("xlsx has exactly Grupo and Vista docente sheets", wb.sheetnames == ["Grupo", "Vista docente"], {"sheetnames": wb.sheetnames})
            ws = wb["Vista docente"]
            expected_cols = ["Cedula", "Nombre", "Apellidos", "Programa", "Promedio", "Estado matricula", "Flags de vulnerabilidad", "N flags"]
            header, export_rows = rows_as_dicts(ws)
            check("Vista docente has exactly 8 expected columns", header == expected_cols, {"header": header, "max_column": ws.max_column})
            check("PREEAC2602B010224 has 85 student rows + 1 header", ws.max_row == 86 and len(export_rows) == 85, {"max_row": ws.max_row, "data_rows": len(export_rows)})
            example_flags = [r for r in export_rows if r.get("Flags de vulnerabilidad") == "SISBEN A5 | Rural"]
            check("flags column consolidates visible flags example SISBEN A5 | Rural", len(example_flags) > 0, {"matching_rows": len(example_flags), "sample_flags": sorted(set(str(r.get("Flags de vulnerabilidad")) for r in export_rows if r.get("Flags de vulnerabilidad")))[:10]})

            detail_resp = requests.get(f"{API}/admin/grupos/{GRUPO_TARGET}", headers=headers(super_token), timeout=30)
            check("group detail API available for UI comparison", detail_resp.status_code == 200, {"status": detail_resp.status_code})
            if detail_resp.status_code == 200:
                detail = detail_resp.json()
                check("export row count matches detail visible matriculados", len(export_rows) == len(detail.get("estudiantes", [])) == detail.get("total_estudiantes"), {"export_rows": len(export_rows), "detail_estudiantes": len(detail.get("estudiantes", [])), "detail_total": detail.get("total_estudiantes")})
                export_by_ced = {str(r.get("Cedula")): r for r in export_rows}
                mismatches = []
                for s in detail.get("estudiantes", [])[:85]:
                    ced = str(s.get("cedula"))
                    expected_flag, expected_n = ui_flags(s)
                    er = export_by_ced.get(ced)
                    if not er or er.get("Flags de vulnerabilidad") != expected_flag or int(er.get("N flags") or 0) != expected_n:
                        mismatches.append({"cedula": ced, "expected": expected_flag, "expected_n": expected_n, "actual": er})
                check("export flags match the detail-view visible flag logic for all rows", len(mismatches) == 0, {"mismatch_count": len(mismatches), "first_mismatch": mismatches[:1]})

        # Docente without download_enabled=true should be forbidden by the new endpoint.
        docente = db.users.find_one({"email": "docente.demo@iudigital.edu.co"}, {"_id": 0, "id": 1, "download_enabled": 1, "role": 1})
        settings = db.system_settings.find_one({"_id": "global"}, {"_id": 0}) or {}
        docente_login = login("docente.demo@iudigital.edu.co", "Docente2026")
        check("docente demo login", docente_login.status_code == 200, {"status": docente_login.status_code, "download_enabled_db": (docente or {}).get("download_enabled"), "global_download": settings.get("docente_downloads_globally_enabled", False)})
        if docente_login.status_code == 200:
            docente_token = docente_login.json()["access_token"]
            own_group = None
            if docente:
                own_group = find_group_with_students(db, {"docente_id": docente["id"]})
            test_group = (own_group or {"codigo_grupo": GRUPO_TARGET})["codigo_grupo"]
            r_doc = requests.get(f"{API}/exports/grupo/{test_group}/vista?fmt=xlsx", headers=headers(docente_token), timeout=30)
            detail_text = r_doc.text[:200] if r_doc.headers.get("content-type", "").startswith("application/json") else "binary/non-json"
            check("docente without download_enabled=true receives 403 on vista endpoint", (docente or {}).get("download_enabled") is not True and r_doc.status_code == 403, {"status": r_doc.status_code, "body": detail_text, "tested_group": test_group, "own_group": bool(own_group)})

        # Scope enforcement: temporarily enable downloads for decano/coordinador if disabled, then restore.
        for email in ["decano.test@iudigital.edu.co", "coord.test@iudigital.edu.co"]:
            u = db.users.find_one({"email": email}, {"_id": 0, "id": 1, "email": 1, "role": 1, "download_enabled": 1, "facultad_id": 1, "programa_id": 1})
            if u:
                original_user_values[email] = u.get("download_enabled", None)
                if u.get("download_enabled") is not True:
                    db.users.update_one({"id": u["id"]}, {"$set": {"download_enabled": True}})

        decano_user = db.users.find_one({"email": "decano.test@iudigital.edu.co"}, {"_id": 0})
        coord_user = db.users.find_one({"email": "coord.test@iudigital.edu.co"}, {"_id": 0})
        fac = db.facultades.find_one({"id": decano_user.get("facultad_id")}, {"_id": 0, "nombre": 1}) if decano_user else None
        prog = db.programas.find_one({"id": coord_user.get("programa_id")}, {"_id": 0, "nombre": 1}) if coord_user else None
        decano_fac = (fac or {}).get("nombre")
        coord_prog = (prog or {}).get("nombre")

        scope_cases = []
        if decano_fac:
            g_in = find_group_by_predicate(db, lambda g: norm(g.get("facultad")) == norm(decano_fac))
            g_out = find_group_by_predicate(db, lambda g: norm(g.get("facultad")) != norm(decano_fac))
            scope_cases.append(("decano", "decano.test@iudigital.edu.co", "Decano2026!", g_in, g_out, decano_fac))
        if coord_prog:
            g_in = find_group_by_predicate(db, lambda g: norm(g.get("programa")) == norm(coord_prog))
            g_out = find_group_by_predicate(db, lambda g: norm(g.get("programa")) != norm(coord_prog))
            scope_cases.append(("coordinador", "coord.test@iudigital.edu.co", "Coord2026!", g_in, g_out, coord_prog))

        for role_name, email, password, g_in, g_out, scope_name in scope_cases:
            lr = login(email, password)
            check(f"{role_name} login for scope test", lr.status_code == 200, {"status": lr.status_code, "scope": scope_name})
            if lr.status_code != 200:
                continue
            tok = lr.json()["access_token"]
            status_in = None
            status_out = None
            if g_in:
                status_in = requests.get(f"{API}/exports/grupo/{g_in['codigo_grupo']}/vista?fmt=xlsx", headers=headers(tok), timeout=30).status_code
            if g_out:
                resp_out = requests.get(f"{API}/exports/grupo/{g_out['codigo_grupo']}/vista?fmt=xlsx", headers=headers(tok), timeout=30)
                status_out = resp_out.status_code
            results["scope_test_groups"][role_name] = {"scope": scope_name, "inside": g_in, "outside": g_out, "inside_status": status_in, "outside_status": status_out}
            check(f"{role_name} can download group inside scope", status_in == 200, {"inside_group": g_in, "status": status_in})
            check(f"{role_name} is forbidden from downloading group outside scope", status_out == 403, {"outside_group": g_out, "status": status_out})

    finally:
        for email, original_download in original_user_values.items():
            if original_download is None:
                db.users.update_one({"email": email}, {"$unset": {"download_enabled": ""}})
            else:
                db.users.update_one({"email": email}, {"$set": {"download_enabled": original_download}})
            results["restored_users"].append({"email": email, "download_enabled": original_download})
        client.close()

    print(json.dumps(results, ensure_ascii=False, indent=2, default=str))
    return 1 if results["failures"] else 0


if __name__ == "__main__":
    raise SystemExit(main())